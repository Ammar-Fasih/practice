import openpyxl
import time
import os
import req_functions as fn
import datetime
import shutil

t1=time.time()


def WSPrefix(ws,table,col_index,prefix):
    for index,row in enumerate(ws[table.ref]):
        if index !=0:
            cell = row[col_index]
            new_value = prefix + cell.value
            cell.value = new_value

def SLPrefix(ws,table,col_index,prefix):
    for index,row in enumerate(ws[table.ref]):
        if index !=0:
            cell = row[col_index]
            new_value = prefix + cell.value[1:]
            cell.value = new_value

os.makedirs('output_files',exist_ok=True)

def main():

    bucket_name = 'hrp_piip_bucket'

    input_file = 'interProcess_files/wholesaleFile.xlsx'   

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    table_ws = ws.tables['Table_UC_Wholesale']
    col_index_ws = table_ws.column_names.index('Base Model')
    prefixes = ['B','C','H','L']
    filenames= ['UC-WholeSale-Bohn.xlsx','UC-WholeSale-ClimateControl.xlsx','UC-WholeSale-Chandler.xlsx','UC-WholeSale-Larkin.xlsx']

    for prefix, filename in zip(prefixes, filenames):
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        table_ws = ws.tables['Table_UC_Wholesale']
        col_index_ws = table_ws.column_names.index('Base Model')
        WSPrefix(ws, table_ws, col_index_ws, prefix)

        ws = wb['Shipped Loose']
        table_sl1 = ws.tables['Table_medium_profile_drain_pan_kits']
        col_index_sl1 = table_sl1.column_names.index('Models')
        SLPrefix(ws, table_sl1, col_index_sl1, prefix)

        ws = wb['Shipped Loose']
        table_sl1 = ws.tables['Table_medium_profile_electric_pan_kits']
        col_index_sl1 = table_sl1.column_names.index('Models')
        SLPrefix(ws, table_sl1, col_index_sl1, prefix)
        wb.save(f'output_files/{filename}')

    print('04 Brands files created')

    date = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    output_folder = f'output_files_{date}'
    fn.create_folder(bucket_name, output_folder)
    
    files_to_upload = os.listdir('output_files')
    for file in files_to_upload:
        fn.upload_file(bucket_name, f'output_files/{file}', f'{output_folder}/{file}')
    
    shutil.rmtree('input_files')
    shutil.rmtree('interProcess_files')

    print(f'files are stored in GCS bucket: {bucket_name} - folder: {output_folder}')
    print(f'total time: {time.time()-t1}')

if __name__ == "__main__":
    main()