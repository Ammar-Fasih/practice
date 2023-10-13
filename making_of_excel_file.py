import os
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, column_index_from_string
import time




# to autofit cols as per their length
def autofit_columns(worksheet,col_start,col_end):
    # Dictionary to store the maximum content length for each column
    max_length = {}

    col_start = column_index_from_string(col_start)
    col_end = column_index_from_string(col_end)
    # Iterate through all rows and columns in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=col_start, max_col=col_end):
        for cell in row:
            # Get the column letter (e.g., 'A', 'B', 'C')
            column_letter = get_column_letter(cell.column)

            # Calculate the length of the cell's content
            cell_length = len(str(cell.value))
            # Update max_length for the column if necessary
            if column_letter not in max_length or cell_length > max_length[column_letter]:
                max_length[column_letter] = cell_length
    # Set the column widths based on the maximum content length
    for column_letter, length in max_length.items():
        worksheet.column_dimensions[column_letter].width = length

# to place the multiline text in consecutive top-bottom cells
def wrap_alt(worksheet,text_list,column):
    text_list = text_list.split('\n')
    for foot in text_list:
        worksheet.append([foot])
        worksheet.merge_cells(f"A{worksheet.max_row}:{get_column_letter(column)}{worksheet.max_row}")

# main function to read csv and create tables in xlsx
def shipped_loose_file(worksheet,csv_files,mapping_dict,header_style_01,sub_heading_style):
    # Create a list to store the tables
    tables = []

    # Import data from CSV files and create tables
    for csv_file in csv_files:

        # read csv as panda df
        df = pd.read_csv(csv_file)

        # get dimension of df
        row,column = df.shape
        # reference of table to generate - we have added 3 because we are adding 3 rows before each table
        ref = f"A{worksheet.max_row+3}:{get_column_letter(column)}{worksheet.max_row+3+row}"

        # get the csv name from the whole path
        csv_file = csv_file.split("/")[-1]
        csv_file = csv_file.replace(' ','_')

        # creating table
        table = Table(displayName=f"Table_{csv_file[:-4]}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium16", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)

        # title of table on sheet - getting from the mapping_dict
        worksheet.append([mapping_dict[csv_file.replace('.csv','')][0]])
        # setting the header row style like font, color, size
        worksheet[f'A{worksheet.max_row}'].style = header_style_01
        # adjusting the height of the row as 50
        worksheet.row_dimensions[int(worksheet.max_row)].height = 50
        # merging cell for title as exact 6 columns to match all
        worksheet.merge_cells(f"A{worksheet.max_row}:{get_column_letter(6)}{worksheet.max_row}")
        
        # get and writing the sub_heading
        sub_heading = mapping_dict[csv_file.replace('.csv','')][1]
        worksheet.append([sub_heading])
        worksheet[f'A{worksheet.max_row}'].style = sub_heading_style
        # setting the height of the sub heading row to be 50
        worksheet.row_dimensions[int(worksheet.max_row)].height = 50
        # merging the cells to exact 6 columns
        worksheet.merge_cells(f"A{worksheet.max_row}:{get_column_letter(6)}{worksheet.max_row}")

        # define headers of table and write them
        worksheet.append(list(df.columns))
        
        # write data to sheet by each row
        for row in df.itertuples(index=False):
            worksheet.append(list(row))
        tables.append(table)

    # Add the tables to the worksheet
    for table in tables:
        worksheet.add_table(table)

# to wrap the cols (not in use)
def wrap_columns (worksheet,col_start,col_end):
    col_start = column_index_from_string(col_start)
    col_end = column_index_from_string(col_end)
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=col_start, max_col=col_end):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

# to find col width (not in use - not working properly)
def col_width(worksheet,col_name,width):
    worksheet.column_dimensions[col_name].width = width

def uc_file(worksheet,df):

    # get dimension of df
    row,column = df.shape
    # reference of table to generate
    ref = f"A{worksheet.max_row+1}:{get_column_letter(column)}{worksheet.max_row+1+row}"

    table = Table(displayName=f"Table_UC_Wholesale", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium16", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)

    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False):
            worksheet.append(list(row))

def custom_sort(path):

    def sub_sort(path):
        if 'shipped_loose' in path:
            return 0
        elif 'expansion_valve' in path:
            return 1
        elif 'liquid_line_solenoid_valves' in path:
            return 2
        else:
            return 3
    if 'low_profile' in path:
        return (0,sub_sort(path),path)
    elif 'large_unit_coolers' in path:
        return (1,sub_sort(path),path)
    elif 'low_velocity_center_mount' in path:
        return (2,sub_sort(path),path)
    elif 'center_mount' in path:
        return (3,sub_sort(path),path)
    elif 'medium_profile' in path:
        return (4,sub_sort(path),path)
    else:
        return (5,path)

def main(eff_date):

    t1= time.time()
    
    # define variables
    eff_date = eff_date
    headingmapping = 'footer_mapping_shippedLooseCU.csv'
    sorted_folder = 'input'
    # wholesale_input = 'interProcess_files/uc_wholesaler.csv'
    outputFile = 'wholesaleFile.xlsx'

    df_headingMapping = pd.read_csv(headingmapping)

    # create dict for footer mapping
    tables = df_headingMapping['table'].tolist()
    heading = df_headingMapping['heading'].tolist()
    sub_heading = df_headingMapping['sub_heading'].tolist()
    heading_subHeading = [[heading[i], sub_heading[i]] for i in range(len(heading))]
    mapping_dict = dict(zip(tables,heading_subHeading))

    # list all the files in sorted folder
    folders = os.listdir(sorted_folder)
    folders = [f'{sorted_folder}/{folder}' for folder in folders]

    # folders = sorted(folders,key=custom_sort)

    # define styles
    header_style_01 = NamedStyle(name='style_01')
    header_style_01.font = Font(name='Arial', size=18, bold=True, color='000066CC')
    header_style_01.alignment = Alignment(wrap_text=True)

    sub_heading_style = NamedStyle(name="sub_heading_style")
    # sub_heading_style.font = Font(size=9)
    sub_heading_style.alignment = Alignment(wrap_text=True,vertical='top')

    
    # # df_wholesale = pd.read_csv(wholesale_input)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Wholesale'

    worksheet.append([f'Price Effective: {eff_date}'])
    # uc_file(worksheet,df_wholesale)

    workbook.save(outputFile)

    # Create a new workbook
    workbook = openpyxl.load_workbook(outputFile)

    # Select a worksheet
    worksheet = workbook.create_sheet('Shipped Loose')
    new_worksheet = workbook.active

    worksheet.append([f'Price Effective: {eff_date}'])
    shipped_loose_file(worksheet,folders,mapping_dict,header_style_01,sub_heading_style)
    col_width(worksheet,'A',30)
    col_width(worksheet,'B',30)
    col_width(worksheet,'C',30)
    col_width(worksheet,'D',30)
    col_width(worksheet,'E',30)
    col_width(worksheet,'F',30)
    # col_width(worksheet,'G',20)


    # Save the workbook
    workbook.save(outputFile)

    # Close the workbook
    workbook.close()

    print('='*100)
    print('Excel File Created')
    print(f'time consumed: {time.time() - t1}')
    print('='*100)

if __name__ == "__main__":
    main('13-Jan')